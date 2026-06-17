import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data = json.load(open('outputs/worklist-data.json'))
CATN = {'ufos_aliens':'UFOs & Aliens','cryptids':'Cryptids','ghosts_hauntings':'Ghosts & Hauntings',
 'psychic_phenomena':'Psychic Phenomena','consciousness_practices':'Consciousness Practices',
 'psychological_experiences':'Psychological Experiences','religion_mythology':'Religion & Mythology',
 'esoteric_practices':'Esoteric & Occult'}

NAVY='1F3A5F'; LBLUE='DCE6F1'; GREY='F2F2F2'
hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hfill = PatternFill('solid', fgColor=NAVY)
tf = Font(name='Arial', size=10)
bf = Font(name='Arial', bold=True, size=10)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, ncol, row=1):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c); cell.font = hf; cell.fill = hfill
        cell.alignment = center; cell.border = border

def zebra(ws, r, ncol):
    if r % 2 == 0:
        for c in range(1, ncol+1): ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=GREY)

# ---- Summary ----
s = wb.active; s.title = 'Summary'
s['A1'] = 'Paradocs Phenomena — Image & Report Worklist'; s['A1'].font = Font(name='Arial', bold=True, size=15, color=NAVY)
s['A2'] = 'Generated ' + datetime.date.today().isoformat(); s['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
rows = [
 ('', ''),
 ('Active phenomena (visible categories)', 528),
 ('Images applied this round (from Drive sheet)', 428),
 ('NEEDS IMAGES — not yet sourced this round', len(data['needImg'])),
 ('NEEDS REPORTS — active but 0 reports, so invisible until ingested', len(data['needRep'])),
 ('', ''),
]
r = 4
for label, val in rows:
    s.cell(row=r, column=1, value=label).font = bf if val != '' else tf
    if val != '': c = s.cell(row=r, column=2, value=val); c.font = bf; c.alignment = center
    r += 1
notes = [
 'TABS:',
 '  • "Needs Images" — 100 active phenomena whose images were NOT sourced in this round.',
 '       Sorted by category, then by report_count (highest-traffic first). Old/legacy images do not count.',
 '  • "Needs Reports" — 12 active phenomena with zero reports; they do NOT appear on the site',
 '       until reports are ingested for them (mostly Religion & Mythology). Most already have images.',
 '',
 'Site visibility rule: the encyclopedia grid shows active phenomena with report_count > 0 in non-hidden categories.',
 'Deprecated / hidden / archived phenomena are excluded from both lists.',
]
r += 1
for n in notes:
    cell = s.cell(row=r, column=1, value=n); cell.font = Font(name='Arial', size=10, bold=n.endswith(':')); r += 1
s.column_dimensions['A'].width = 80; s.column_dimensions['B'].width = 12

# ---- Needs Images ----
wi = wb.create_sheet('Needs Images')
hdr = ['Category','Phenomenon','Slug','Reports','Visible now','Has old image']
wi.append(hdr); style_header(wi, len(hdr))
for i, p in enumerate(data['needImg'], start=2):
    wi.append([CATN.get(p['category'],p['category']), p['name'], p['slug'], p['report_count'], p['visible_now'], p['has_old_image']])
    for c in range(1,len(hdr)+1):
        cell = wi.cell(row=i, column=c); cell.font = tf; cell.border = border
        cell.alignment = center if c in (3,4,5,6) else left
    zebra(wi, i, len(hdr))
wi.cell(row=1,column=4).comment = None
for col,w in zip('ABCDEF',[24,34,30,10,12,13]): wi.column_dimensions[col].width = w
wi.freeze_panes = 'A2'; wi.auto_filter.ref = 'A1:F' + str(len(data['needImg'])+1)

# ---- Needs Reports ----
wr = wb.create_sheet('Needs Reports')
hdr2 = ['Category','Phenomenon','Slug','Has image','Sourced this round']
wr.append(hdr2); style_header(wr, len(hdr2))
for i, p in enumerate(data['needRep'], start=2):
    wr.append([CATN.get(p['category'],p['category']), p['name'], p['slug'], p['has_image'], p['sourced_this_round']])
    for c in range(1,len(hdr2)+1):
        cell = wr.cell(row=i, column=c); cell.font = tf; cell.border = border
        cell.alignment = center if c in (3,4,5) else left
    zebra(wr, i, len(hdr2))
for col,w in zip('ABCDE',[24,34,30,11,18]): wr.column_dimensions[col].width = w
wr.freeze_panes = 'A2'; wr.auto_filter.ref = 'A1:E' + str(len(data['needRep'])+1)

wb.save('Paradocs_Phenomena_Worklist.xlsx')
print('wrote Paradocs_Phenomena_Worklist.xlsx | needs-images', len(data['needImg']), '| needs-reports', len(data['needRep']))
